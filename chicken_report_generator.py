"""
炸雞對帳報告生成器
生成炸雞品項的對帳報告
"""
import pandas as pd
import os
from datetime import datetime
from typing import Dict, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

logger = logging.getLogger(__name__)

class ChickenReportGenerator:
    """炸雞對帳報告生成器類別"""
    
    def __init__(self, output_dir: str = "chicken_reports"):
        """
        初始化報告生成器
        
        Args:
            output_dir (str): 輸出目錄
        """
        self.output_dir = output_dir
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """確保輸出目錄存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"建立輸出目錄: {self.output_dir}")
    
    def generate_excel_report(self, settlement_report: Dict, filename: str = None) -> str:
        """
        生成 Excel 格式的炸雞對帳報告
        
        Args:
            settlement_report (Dict): 炸雞對帳報告資料
            filename (str): 檔案名稱
            
        Returns:
            str: 檔案路徑
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"炸雞對帳報告_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # 建立工作簿
            wb = Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 建立各個工作表
            self._create_summary_sheet(wb, settlement_report)
            self._create_product_summary_sheet(wb, settlement_report)
            self._create_daily_summary_sheet(wb, settlement_report)
            self._create_settlement_sheet(wb, settlement_report)
            self._create_detail_sheet(wb, settlement_report)
            
            # 儲存檔案
            wb.save(filepath)
            logger.info(f"Excel 報告已生成: {filepath}")
            return filepath
            
        except Exception as error:
            logger.error(f"生成 Excel 報告時發生錯誤: {error}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, report: Dict):
        """建立摘要工作表"""
        ws = wb.create_sheet("炸雞對帳摘要")
        
        # 設定標題
        title_font = Font(name="微軟正黑體", size=16, bold=True)
        header_font = Font(name="微軟正黑體", size=12, bold=True)
        content_font = Font(name="微軟正黑體", size=11)
        
        # 標題
        ws['A1'] = "🍗 炸雞對帳報告"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 期間資訊
        ws['A3'] = "對帳期間:"
        ws['B3'] = report['期間']
        ws['A3'].font = header_font
        ws['B3'].font = content_font
        
        # 摘要資料
        summary_data = [
            ("總銷售金額", f"${report['總銷售金額']:,}"),
            ("總銷售數量", f"{report['總銷售數量']:,} 份"),
            ("總訂單數", f"{report['總訂單數']:,} 筆"),
            ("品項種類", f"{report['品項種類']} 種"),
            ("平均單價", f"${report['平均單價']:.2f}"),
            ("", ""),
            ("🍗 炸雞老闆應付金額", f"${report['炸雞老闆應付金額']:,}"),
            ("成本比例", f"{report['成本比例']*100:.1f}%"),
            ("利潤", f"${report['利潤']:,}")
        ]
        
        for i, (label, value) in enumerate(summary_data, start=5):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = header_font
            ws[f'B{i}'].font = content_font
            
            # 特別標示炸雞老闆應付金額
            if "炸雞老闆應付金額" in label:
                ws[f'A{i}'].font = Font(name="微軟正黑體", size=12, bold=True, color="FF0000")
                ws[f'B{i}'].font = Font(name="微軟正黑體", size=12, bold=True, color="FF0000")
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def _create_product_summary_sheet(self, wb: Workbook, report: Dict):
        """建立品項摘要工作表"""
        ws = wb.create_sheet("品項摘要")
        
        # 標題
        ws['A1'] = "各炸雞品項銷售摘要"
        ws['A1'].font = Font(name="微軟正黑體", size=14, bold=True)
        
        # 品項摘要資料
        if not report['品項摘要'].empty:
            # 標題列
            headers = ['品項', '總數量', '總金額', '平均單價']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(name="微軟正黑體", size=11, bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 資料列
            for row_idx, (_, row) in enumerate(report['品項摘要'].iterrows(), start=4):
                ws.cell(row=row_idx, column=1, value=row['品項'])
                ws.cell(row=row_idx, column=2, value=row['總數量'])
                ws.cell(row=row_idx, column=3, value=row['總金額'])
                ws.cell(row=row_idx, column=4, value=row['平均單價'])
            
            # 建立圖表
            chart = BarChart()
            chart.title = "各品項銷售金額"
            chart.x_axis.title = "品項"
            chart.y_axis.title = "金額"
            
            data = Reference(ws, min_col=3, min_row=3, max_row=3+len(report['品項摘要']))
            categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(report['品項摘要']))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, "F3")
        
        # 調整欄寬
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_daily_summary_sheet(self, wb: Workbook, report: Dict):
        """建立每日摘要工作表"""
        ws = wb.create_sheet("每日摘要")
        
        # 標題
        ws['A1'] = "每日炸雞銷售摘要"
        ws['A1'].font = Font(name="微軟正黑體", size=14, bold=True)
        
        # 每日摘要資料
        if not report['每日摘要'].empty:
            # 標題列
            headers = ['日期', '總數量', '總金額']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(name="微軟正黑體", size=11, bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 資料列
            for row_idx, (_, row) in enumerate(report['每日摘要'].iterrows(), start=4):
                ws.cell(row=row_idx, column=1, value=row['日期'].strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=2, value=row['總數量'])
                ws.cell(row=row_idx, column=3, value=row['總金額'])
            
            # 建立圖表
            chart = BarChart()
            chart.title = "每日銷售金額"
            chart.x_axis.title = "日期"
            chart.y_axis.title = "金額"
            
            data = Reference(ws, min_col=3, min_row=3, max_row=3+len(report['每日摘要']))
            categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(report['每日摘要']))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, "F3")
        
        # 調整欄寬
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 15
    
    def _create_settlement_sheet(self, wb: Workbook, report: Dict):
        """建立對帳工作表"""
        ws = wb.create_sheet("對帳明細")
        
        # 標題
        ws['A1'] = "🍗 炸雞老闆對帳明細"
        ws['A1'].font = Font(name="微軟正黑體", size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 對帳資訊
        settlement_data = [
            ("對帳期間", report['期間']),
            ("總銷售金額", f"${report['總銷售金額']:,}"),
            ("成本比例", f"{report['成本比例']*100:.1f}%"),
            ("", ""),
            ("🍗 炸雞老闆應付金額", f"${report['炸雞老闆應付金額']:,}"),
            ("利潤", f"${report['利潤']:,}"),
            ("", ""),
            ("備註", "此金額為炸雞品項的對帳金額，請確認後付款")
        ]
        
        for i, (label, value) in enumerate(settlement_data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(name="微軟正黑體", size=11, bold=True)
            ws[f'B{i}'].font = Font(name="微軟正黑體", size=11)
            
            # 特別標示應付金額
            if "炸雞老闆應付金額" in label:
                ws[f'A{i}'].font = Font(name="微軟正黑體", size=14, bold=True, color="FF0000")
                ws[f'B{i}'].font = Font(name="微軟正黑體", size=14, bold=True, color="FF0000")
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def _create_detail_sheet(self, wb: Workbook, report: Dict):
        """建立詳細資料工作表"""
        ws = wb.create_sheet("詳細資料")
        
        # 標題
        ws['A1'] = "詳細炸雞銷售資料"
        ws['A1'].font = Font(name="微軟正黑體", size=14, bold=True)
        
        # 詳細資料
        if not report['詳細資料'].empty:
            # 標題列
            headers = ['日期', '品項', '數量', '單價', '小計']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(name="微軟正黑體", size=11, bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 資料列
            for row_idx, (_, row) in enumerate(report['詳細資料'].iterrows(), start=4):
                ws.cell(row=row_idx, column=1, value=row['日期'].strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=2, value=row['品項'])
                ws.cell(row=row_idx, column=3, value=row['數量'])
                ws.cell(row=row_idx, column=4, value=row['單價'])
                ws.cell(row=row_idx, column=5, value=row['小計'])
        
        # 調整欄寬
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def generate_text_report(self, settlement_report: Dict) -> str:
        """
        生成文字格式的炸雞對帳報告
        
        Args:
            settlement_report (Dict): 炸雞對帳報告資料
            
        Returns:
            str: 文字報告內容
        """
        try:
            report_lines = []
            report_lines.append("=" * 60)
            report_lines.append("🍗 炸雞對帳報告")
            report_lines.append("=" * 60)
            report_lines.append(f"對帳期間: {settlement_report['期間']}")
            report_lines.append(f"總銷售金額: ${settlement_report['總銷售金額']:,}")
            report_lines.append(f"總銷售數量: {settlement_report['總銷售數量']:,} 份")
            report_lines.append(f"總訂單數: {settlement_report['總訂單數']:,} 筆")
            report_lines.append(f"品項種類: {settlement_report['品項種類']} 種")
            report_lines.append(f"平均單價: ${settlement_report['平均單價']:.2f}")
            report_lines.append("")
            report_lines.append("🍗 炸雞老闆對帳明細:")
            report_lines.append("-" * 40)
            report_lines.append(f"炸雞老闆應付金額: ${settlement_report['炸雞老闆應付金額']:,}")
            report_lines.append(f"成本比例: {settlement_report['成本比例']*100:.1f}%")
            report_lines.append(f"利潤: ${settlement_report['利潤']:,}")
            report_lines.append("")
            
            # 品項摘要
            if not settlement_report['品項摘要'].empty:
                report_lines.append("各品項銷售摘要:")
                report_lines.append("-" * 40)
                for _, row in settlement_report['品項摘要'].iterrows():
                    report_lines.append(f"{row['品項']}: {row['總數量']} 份, ${row['總金額']:,}")
                report_lines.append("")
            
            # 每日摘要
            if not settlement_report['每日摘要'].empty:
                report_lines.append("每日銷售摘要:")
                report_lines.append("-" * 40)
                for _, row in settlement_report['每日摘要'].iterrows():
                    report_lines.append(f"{row['日期'].strftime('%Y-%m-%d')}: {row['總數量']} 份, ${row['總金額']:,}")
            
            report_content = "\n".join(report_lines)
            logger.info("文字報告生成完成")
            return report_content
            
        except Exception as error:
            logger.error(f"生成文字報告時發生錯誤: {error}")
            raise
